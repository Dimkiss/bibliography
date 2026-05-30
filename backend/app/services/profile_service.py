"""
Сервис профиля сотрудника.

Публикации, статистика и Excel-отчёт для текущего пользователя
(фильтрация по author_id из связки users → authors).
"""

from __future__ import annotations

import io
import math
from typing import Any

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.services.articles.bibliographic_reference import (
    build_bibliographic_reference,
    build_pages_fallback_select_sql,
)
from app.services.articles import pdf_files
from app.schemas.article import ArticleListItem, ArticleListResponse, PaginationMeta

DEFAULT_PAGE_SIZE = 20
MAX_PAGE_SIZE = 100

SOURCE_TITLE_EXPR = """
    COALESCE(
        NULLIF(jn.JournalName, ''),
        NULLIF(j.jname, ''),
        NULLIF(a.Title_of_Material_F9, ''),
        NULLIF(a.Edition_F15, '')
    )
"""


def _extract_databases(row: dict[str, Any]) -> list[str]:
    items: list[str] = []
    if row.get("wos_flag"):
        items.append("Web of Science")
    if row.get("scopus_flag"):
        items.append("Scopus")
    if row.get("white_list_flag"):
        items.append("Белый список")
    if row.get("rinc_flag"):
        items.append("РИНЦ")
    if row.get("rinc_core_flag"):
        items.append("Ядро РИНЦ")
    if row.get("rsci_flag"):
        items.append("RSCI")
    if row.get("vak_flag"):
        items.append("ВАК")
    return items


def _parse_csv_list(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split("|||") if item.strip()]


PUBLICATION_TYPE_CATEGORY_FLAGS: dict[str, tuple[str, ...]] = {
    "article": ("ST",),
    "conference": ("MA", "DO", "PD", "SD", "TE", "TR"),
    "monograph": ("MO",),
    "chapter": ("GL",),
    "patent_method": ("PA", "MP", "AS", "LI"),
}

PUBLICATION_TYPE_PRIMARY_FLAGS: tuple[str, ...] = tuple(
    flag
    for flags in PUBLICATION_TYPE_CATEGORY_FLAGS.values()
    for flag in flags
)


def _normalize_str_list(values: list[str] | None) -> list[str]:
    if not values:
        return []

    return [value.strip() for value in values if value.strip()]


def _build_in_clause(prefix: str, values: list[str], params: dict[str, Any]) -> str:
    placeholders: list[str] = []

    for index, value in enumerate(values):
        key = f"{prefix}_{index}"
        placeholders.append(f":{key}")
        params[key] = value

    return ", ".join(placeholders)


def _build_publication_search_filter(
    text_query: str | None,
    params: dict[str, Any],
) -> str:
    if not text_query or not text_query.strip():
        return ""

    params["text_query"] = f"%{text_query.strip()}%"
    return f"""
        AND (
            a.Title_Analitic_F4 LIKE :text_query
            OR a.Author_Analitic_F1 LIKE :text_query
            OR a.DOI LIKE :text_query
            OR {SOURCE_TITLE_EXPR} LIKE :text_query
            OR EXISTS (
                SELECT 1
                FROM articlehasauthor aha_search
                JOIN authors au_search
                    ON au_search.authorID = aha_search.authorID_f
                WHERE aha_search.Record_ID_f = a.Record_ID
                  AND au_search.authorName LIKE :text_query
            )
            OR EXISTS (
                SELECT 1
                FROM articlehaskeywords ahk_search
                JOIN keywords k_search
                    ON k_search.K_ID = ahk_search.Keyword_ID_f
                WHERE ahk_search.Record_ID_f = a.Record_ID
                  AND k_search.Keyword LIKE :text_query
            )
        )
    """


def _build_publication_type_filter(
    publication_types: list[str],
    params: dict[str, Any],
) -> str:
    if not publication_types:
        return ""

    category_conditions: list[str] = []
    direct_values: list[str] = []

    for publication_type in publication_types:
        flags = PUBLICATION_TYPE_CATEGORY_FLAGS.get(publication_type)
        if flags:
            in_clause = _build_in_clause(
                f"pub_type_{publication_type}",
                list(flags),
                params,
            )
            category_conditions.append(f"aht.TypeOfPublication_f IN ({in_clause})")
        elif publication_type == "other":
            in_clause = _build_in_clause(
                "pub_type_primary",
                list(PUBLICATION_TYPE_PRIMARY_FLAGS),
                params,
            )
            category_conditions.append(f"aht.TypeOfPublication_f NOT IN ({in_clause})")
        else:
            direct_values.append(publication_type)

    if direct_values:
        in_clause = _build_in_clause("pub_type_direct", direct_values, params)
        category_conditions.append(
            f"(top.TOP_Flag IN ({in_clause}) OR top.TOP_Name IN ({in_clause}))"
        )

    if not category_conditions:
        return ""

    return (
        """
        AND EXISTS (
            SELECT 1
            FROM articlehastop aht
            JOIN typesofpublications top
                ON top.TOP_Flag = aht.TypeOfPublication_f
            WHERE aht.Record_ID_f = a.Record_ID
              AND (
        """
        + " OR ".join(f"({condition})" for condition in category_conditions)
        + """
              )
        )
        """
    )


def _build_database_filter(databases: list[str]) -> str:
    if not databases:
        return ""

    conditions: list[str] = []
    for database in databases:
        value = database.lower()
        if value == "wos":
            conditions.append("COALESCE(j.WOS, 0) = 1")
        elif value == "scopus":
            conditions.append("COALESCE(j.Scopus, 0) = 1")
        elif value == "white_list":
            conditions.append("COALESCE(j.LWL, 0) > 0")
        elif value == "rinc":
            conditions.append("COALESCE(j.Rints, 0) = 1")
        elif value == "rinc_core":
            conditions.append("COALESCE(j.RintsCore, 0) = 1")
        elif value == "rsci":
            conditions.append("COALESCE(j.RSCI, 0) = 1")
        elif value == "vak":
            conditions.append("COALESCE(j.BAK, 0) = 1")

    return f"AND ({' OR '.join(conditions)})" if conditions else ""


def _build_original_translation_filter(original_translation_mode: str) -> str:
    if original_translation_mode == "original_only":
        return """
        AND NOT (
            jaa.OriginalVer_ID_f IS NOT NULL
            OR EXISTS (
                SELECT 1
                FROM journalarticlesattributes relation
                WHERE relation.PerVer_ID_f = a.Record_ID
            )
        )
        """

    if original_translation_mode == "translation_only":
        return """
        AND NOT (
            jaa.PerVer_ID_f IS NOT NULL
            OR EXISTS (
                SELECT 1
                FROM journalarticlesattributes relation
                WHERE relation.OriginalVer_ID_f = a.Record_ID
            )
        )
        """

    return ""


# ---------------------------------------------------------------------------
# Список публикаций автора
# ---------------------------------------------------------------------------

def get_profile_publications(
    db: Session,
    author_id: int,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
    year_from: int | None = None,
    year_to: int | None = None,
    text_query: str | None = None,
    publication_types: list[str] | None = None,
    databases: list[str] | None = None,
    original_translation_mode: str = "all",
    sort_by: str = "year",
    sort_order: str = "desc",
) -> ArticleListResponse:
    """Возвращает пагинированный список публикаций для автора."""

    publication_types = _normalize_str_list(publication_types)
    databases = _normalize_str_list(databases)
    sort_dir = "ASC" if sort_order == "asc" else "DESC"
    sort_expr_map = {
        "year": "a.Date_of_Publication_F20",
        "title": "a.Title_Analitic_F4",
        "journal": SOURCE_TITLE_EXPR,
        "quartile": "COALESCE(NULLIF(j.Quartile, ''), NULLIF(j.QuartileScopus, ''))",
    }
    sort_expr = sort_expr_map.get(sort_by, "a.Date_of_Publication_F20")

    year_filter = ""
    params: dict[str, Any] = {"author_id": author_id}

    if year_from is not None:
        year_filter += " AND a.Date_of_Publication_F20 >= :year_from"
        params["year_from"] = year_from
    if year_to is not None:
        year_filter += " AND a.Date_of_Publication_F20 <= :year_to"
        params["year_to"] = year_to

    search_filter = _build_publication_search_filter(text_query, params)
    publication_type_filter = _build_publication_type_filter(publication_types, params)
    database_filter = _build_database_filter(databases)
    original_translation_filter = _build_original_translation_filter(
        original_translation_mode,
    )

    # Считаем total
    count_sql = text(f"""
        SELECT COUNT(DISTINCT a.Record_ID) AS total
        FROM articles a
        JOIN articlehasauthor aha ON aha.Record_ID_f = a.Record_ID
            AND aha.authorID_f = :author_id
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        LEFT JOIN journalnames jn ON jn.JN_ID = j.JN_ID_f
        LEFT JOIN journalarticlesattributes jaa ON jaa.Record_ID_f = a.Record_ID
        WHERE 1 = 1
        {year_filter}
        {search_filter}
        {publication_type_filter}
        {database_filter}
        {original_translation_filter}
    """)
    total = int(db.execute(count_sql, params).scalar() or 0)

    offset = (page - 1) * page_size
    page_params = dict(params)
    page_params["limit"] = page_size
    page_params["offset"] = offset

    # ID-запрос с сортировкой
    id_sql = text(f"""
        SELECT a.Record_ID AS id
        FROM articles a
        JOIN articlehasauthor aha ON aha.Record_ID_f = a.Record_ID
            AND aha.authorID_f = :author_id
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        LEFT JOIN journalnames jn ON jn.JN_ID = j.JN_ID_f
        LEFT JOIN journalarticlesattributes jaa ON jaa.Record_ID_f = a.Record_ID
        WHERE 1 = 1
        {year_filter}
        {search_filter}
        {publication_type_filter}
        {database_filter}
        {original_translation_filter}
        ORDER BY {sort_expr} {sort_dir}, a.Record_ID DESC
        LIMIT :limit OFFSET :offset
    """)
    id_rows = db.execute(id_sql, page_params).mappings().all()
    article_ids = [int(row["id"]) for row in id_rows]

    if not article_ids:
        total_pages = math.ceil(total / page_size) if total > 0 else 0
        return ArticleListResponse(
            items=[],
            pagination=PaginationMeta(
                page=page,
                page_size=page_size,
                total=total,
                total_pages=total_pages,
            ),
        )

    id_placeholders = ", ".join(f":aid_{i}" for i in range(len(article_ids)))
    data_params: dict[str, Any] = {f"aid_{i}": aid for i, aid in enumerate(article_ids)}
    pages_fallback_sql = build_pages_fallback_select_sql(db)

    data_sql = text(f"""
        SELECT
            a.Record_ID AS id,
            a.Record_ID AS record_id,
            a.WorkFormType_f AS work_form_type,
            a.Author_Analitic_F1 AS author_analitic,
            a.Title_Analitic_F4 AS title,
            a.Title_Analitic_F4 AS title_analitic,
            a.Author_of_Material_F7 AS author_of_material,
            a.Title_of_Material_F9 AS title_of_material,
            a.DateOfMeeting_F12 AS date_of_meeting,
            a.Edition_F15 AS edition,
            a.Date_of_Publication_F20 AS date_of_publication,
            a.VolumeID_F22 AS volume,
            a.IssueID_F24 AS issue,
            a.Pages_F25 AS pages,
            {pages_fallback_sql} AS pages_fallback,
            a.ExtentOfWork_F26 AS extent_of_work,
            a.ISBN_F41 AS isbn,
            COALESCE(
                NULLIF(a.Author_Analitic_F1, ''),
                (
                    SELECT GROUP_CONCAT(DISTINCT au.authorName ORDER BY au.authorName SEPARATOR ', ')
                    FROM articlehasauthor aha2
                    JOIN authors au ON au.authorID = aha2.authorID_f
                    WHERE aha2.Record_ID_f = a.Record_ID
                )
            ) AS authors,
            NULLIF(jn.JournalName, '') AS journal_name,
            {SOURCE_TITLE_EXPR} AS journal,
            a.Date_of_Publication_F20 AS year,
            a.DOI AS doi,
            pp.PlaceName AS place_name,
            pn.PublisherName AS publisher_name,
            NULLIF(j.Quartile, '') AS quartile,
            NULLIF(j.QuartileScopus, '') AS quartile_scopus,
            j.Impact_Factor AS impact_factor,
            COALESCE(j.WOS, 0) AS wos_flag,
            COALESCE(j.Scopus, 0) AS scopus_flag,
            COALESCE(j.LWL, 0) AS white_list_flag,
            COALESCE(j.Rints, 0) AS rinc_flag,
            COALESCE(j.RintsCore, 0) AS rinc_core_flag,
            COALESCE(j.RSCI, 0) AS rsci_flag,
            COALESCE(j.BAK, 0) AS vak_flag,
            (
                SELECT GROUP_CONCAT(DISTINCT top.TOP_Name ORDER BY top.TOP_Name SEPARATOR '|||')
                FROM articlehastop aht
                JOIN typesofpublications top ON top.TOP_Flag = aht.TypeOfPublication_f
                WHERE aht.Record_ID_f = a.Record_ID
            ) AS publication_types_csv,
            CASE
                WHEN jaa.OriginalVer_ID_f IS NOT NULL
                    OR EXISTS (
                        SELECT 1 FROM journalarticlesattributes r
                        WHERE r.PerVer_ID_f = a.Record_ID
                    )
                THEN 'translation'
                WHEN jaa.PerVer_ID_f IS NOT NULL
                    OR EXISTS (
                        SELECT 1 FROM journalarticlesattributes r
                        WHERE r.OriginalVer_ID_f = a.Record_ID
                    )
                THEN 'original'
                ELSE NULL
            END AS original_translation
        FROM articles a
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        LEFT JOIN journalnames jn ON jn.JN_ID = j.JN_ID_f
        LEFT JOIN journalarticlesattributes jaa ON jaa.Record_ID_f = a.Record_ID
        LEFT JOIN places pp ON pp.P_ID = a.PlaceOfPublication_F18_f
        LEFT JOIN publishernames pn ON pn.PN_ID = a.PublisherName_F19_f
        WHERE a.Record_ID IN ({id_placeholders})
        ORDER BY FIELD(a.Record_ID, {id_placeholders})
    """)

    rows = db.execute(data_sql, data_params).mappings().all()

    items: list[ArticleListItem] = []
    for row in rows:
        rd = dict(row)
        items.append(
            ArticleListItem(
                id=rd["id"],
                title=rd.get("title"),
                authors=rd.get("authors"),
                journal=rd.get("journal"),
                year=rd.get("year"),
                doi=rd.get("doi"),
                bibliographic_reference=build_bibliographic_reference(rd),
                quartile=rd.get("quartile"),
                quartile_scopus=rd.get("quartile_scopus"),
                publication_types=_parse_csv_list(rd.get("publication_types_csv")),
                databases=_extract_databases(rd),
                original_translation=rd.get("original_translation"),
                has_pdf=pdf_files.article_pdf_exists(rd["id"]),
            )
        )

    total_pages = math.ceil(total / page_size) if total > 0 else 0
    return ArticleListResponse(
        items=items,
        pagination=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=total_pages,
        ),
    )


# ---------------------------------------------------------------------------
# Агрегированная статистика по автору
# ---------------------------------------------------------------------------

def get_profile_stats(
    db: Session,
    author_id: int,
    year_from: int | None = None,
    year_to: int | None = None,
    text_query: str | None = None,
    publication_types: list[str] | None = None,
    databases: list[str] | None = None,
    original_translation_mode: str = "all",
) -> dict[str, Any]:
    """Считает ключевые показатели автора для отображения в шапке профиля."""

    publication_types = _normalize_str_list(publication_types)
    databases = _normalize_str_list(databases)
    year_filter = ""
    params: dict[str, Any] = {"author_id": author_id}
    if year_from is not None:
        year_filter += " AND a.Date_of_Publication_F20 >= :year_from"
        params["year_from"] = year_from
    if year_to is not None:
        year_filter += " AND a.Date_of_Publication_F20 <= :year_to"
        params["year_to"] = year_to

    search_filter = _build_publication_search_filter(text_query, params)
    publication_type_filter = _build_publication_type_filter(publication_types, params)
    database_filter = _build_database_filter(databases)
    original_translation_filter = _build_original_translation_filter(
        original_translation_mode,
    )

    sql = text(f"""
        SELECT
            COUNT(DISTINCT a.Record_ID) AS total,
            SUM(CASE WHEN COALESCE(j.WOS, 0) = 1 THEN 1 ELSE 0 END) AS wos_count,
            SUM(CASE WHEN COALESCE(j.Scopus, 0) = 1 THEN 1 ELSE 0 END) AS scopus_count,
            SUM(CASE WHEN COALESCE(j.BAK, 0) = 1 THEN 1 ELSE 0 END) AS vak_count,
            SUM(CASE WHEN COALESCE(j.LWL, 0) > 0 THEN 1 ELSE 0 END) AS white_list_count,
            ROUND(SUM(COALESCE(j.Impact_Factor, 0)), 3) AS if_total
        FROM articles a
        JOIN articlehasauthor aha ON aha.Record_ID_f = a.Record_ID
            AND aha.authorID_f = :author_id
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        LEFT JOIN journalnames jn ON jn.JN_ID = j.JN_ID_f
        LEFT JOIN journalarticlesattributes jaa ON jaa.Record_ID_f = a.Record_ID
        WHERE 1 = 1
        {year_filter}
        {search_filter}
        {publication_type_filter}
        {database_filter}
        {original_translation_filter}
    """)

    row = db.execute(sql, params).mappings().first()
    if row is None:
        return {"total": 0, "wos_count": 0, "scopus_count": 0,
                "vak_count": 0, "white_list_count": 0, "if_total": 0.0}

    return {
        "total": int(row["total"] or 0),
        "wos_count": int(row["wos_count"] or 0),
        "scopus_count": int(row["scopus_count"] or 0),
        "vak_count": int(row["vak_count"] or 0),
        "white_list_count": int(row["white_list_count"] or 0),
        "if_total": float(row["if_total"] or 0.0),
    }


# ---------------------------------------------------------------------------
# Генерация Excel-отчёта
# ---------------------------------------------------------------------------

def generate_profile_report(
    db: Session,
    author_id: int,
    author_name: str,
    year_from: int | None = None,
    year_to: int | None = None,
    article_ids: list[int] | None = None,
) -> bytes:
    """Формирует xlsx-отчёт по публикациям автора.

    Если передан article_ids — выгружает только указанные статьи
    (с проверкой принадлежности автору через JOIN).
    Иначе — все публикации автора с опциональным фильтром по годам.
    """

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    params: dict[str, Any] = {"author_id": author_id}

    if article_ids:
        placeholders = ", ".join(f":aid_{i}" for i in range(len(article_ids)))
        article_filter = f"AND a.Record_ID IN ({placeholders})"
        for i, aid in enumerate(article_ids):
            params[f"aid_{i}"] = aid
        row_filter = article_filter
    else:
        row_filter = ""
        if year_from is not None:
            row_filter += " AND a.Date_of_Publication_F20 >= :year_from"
            params["year_from"] = year_from
        if year_to is not None:
            row_filter += " AND a.Date_of_Publication_F20 <= :year_to"
            params["year_to"] = year_to

    sql = text(f"""
        SELECT
            a.Record_ID,
            a.WorkFormType_f         AS work_form_type,
            a.Author_Analitic_F1     AS author_analitic,
            a.Title_Analitic_F4      AS title,
            a.Title_Analitic_F4      AS title_analitic,
            a.Author_of_Material_F7  AS author_of_material,
            a.Title_of_Material_F9   AS title_of_material,
            a.DateOfMeeting_F12      AS date_of_meeting,
            a.Edition_F15            AS edition,
            a.Date_of_Publication_F20 AS date_of_publication,
            a.Date_of_Publication_F20 AS year,
            a.VolumeID_F22           AS volume,
            a.IssueID_F24            AS issue,
            a.Pages_F25              AS pages,
            a.ExtentOfWork_F26       AS extent_of_work,
            a.ISBN_F41               AS isbn,
            a.DOI                    AS doi,
            COALESCE(
                NULLIF(a.Author_Analitic_F1, ''),
                (
                    SELECT GROUP_CONCAT(DISTINCT au.authorName ORDER BY au.authorName SEPARATOR ', ')
                    FROM articlehasauthor aha2
                    JOIN authors au ON au.authorID = aha2.authorID_f
                    WHERE aha2.Record_ID_f = a.Record_ID
                )
            ) AS authors,
            NULLIF(jn.JournalName, '') AS journal_name,
            COALESCE(
                NULLIF(jn.JournalName, ''),
                NULLIF(j.jname, ''),
                NULLIF(a.Title_of_Material_F9, ''),
                NULLIF(a.Edition_F15, '')
            ) AS journal,
            jn.ISSN                  AS issn,
            pp.PlaceName             AS place_name,
            pn.PublisherName         AS publisher_name,
            NULLIF(j.Quartile, '')   AS quartile,
            NULLIF(j.QuartileScopus, '') AS quartile_scopus,
            j.Impact_Factor          AS impact_factor,
            COALESCE(j.WOS, 0)       AS wos,
            COALESCE(j.Scopus, 0)    AS scopus,
            COALESCE(j.LWL, 0)       AS white_list_flag,
            j.LWL                    AS white_list_level,
            COALESCE(j.Rints, 0)     AS rinc,
            COALESCE(j.RintsCore, 0) AS rinc_core,
            COALESCE(j.RSCI, 0)      AS rsci,
            COALESCE(j.BAK, 0)       AS vak,
            (
                SELECT GROUP_CONCAT(DISTINCT top.TOP_Name ORDER BY top.TOP_Name SEPARATOR ', ')
                FROM articlehastop aht
                JOIN typesofpublications top ON top.TOP_Flag = aht.TypeOfPublication_f
                WHERE aht.Record_ID_f = a.Record_ID
            ) AS pub_types
        FROM articles a
        JOIN articlehasauthor aha ON aha.Record_ID_f = a.Record_ID
            AND aha.authorID_f = :author_id
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        LEFT JOIN journalnames jn ON jn.JN_ID = j.JN_ID_f
        LEFT JOIN places pp ON pp.P_ID = a.PlaceOfPublication_F18_f
        LEFT JOIN publishernames pn ON pn.PN_ID = a.PublisherName_F19_f
        WHERE 1 = 1
        {row_filter}
        ORDER BY a.Date_of_Publication_F20 DESC, a.Record_ID DESC
    """)
    rows = db.execute(sql, params).mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Публикации"

    # Стили
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="4E5B92")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Заголовок файла
    year_range = ""
    if year_from and year_to:
        year_range = f" ({year_from}–{year_to})"
    elif year_from:
        year_range = f" (с {year_from})"
    elif year_to:
        year_range = f" (по {year_to})"

    # Колонки: (заголовок, ширина)
    headers = [
        ("№",                   5),
        ("Библиографическая\nссылка", 70),
        ("Название",            40),
        ("Авторы",              25),
        ("Журнал / Издание",    25),
        ("ISSN",                12),
        ("Год",                  6),
        ("Том",                  6),
        ("Выпуск",               8),
        ("Страницы",            10),
        ("DOI",                 22),
        ("Тип публикации",      18),
        ("WoS",                  6),
        ("Scopus",               8),
        ("Квартиль WoS",        13),
        ("Квартиль Scopus",     14),
        ("Импакт-\nфактор",     13),
        ("ВАК",                  6),
        ("РИНЦ",                 6),
        ("Ядро\nРИНЦ",          9),
        ("RSCI",                 7),
        ("Белый\nсписок",       10),
        ("Уровень\nБС",          9),
    ]
    TOTAL_COLS = len(headers)

    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    title_cell = ws["A1"]
    title_cell.value = f"Список публикаций: {author_name}{year_range}"
    title_cell.font = Font(name="Calibri", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 40

    # Данные
    for row_idx, row in enumerate(rows, start=3):
        rd = dict(row)

        bib_ref = build_bibliographic_reference(rd)

        wl_level = rd.get("white_list_level")
        wl_quartile = f"Q{wl_level}" if wl_level else ""

        data = [
            row_idx - 2,
            bib_ref,
            rd.get("title") or "",
            rd.get("authors") or "",
            rd.get("journal") or "",
            rd.get("issn") or "",
            rd.get("year"),
            rd.get("volume") or "",
            rd.get("issue") or "",
            rd.get("pages") or "",
            rd.get("doi") or "",
            rd.get("pub_types") or "",
            "Да" if rd.get("wos") else "",
            "Да" if rd.get("scopus") else "",
            rd.get("quartile") or "",
            rd.get("quartile_scopus") or "",
            float(rd["impact_factor"]) if rd.get("impact_factor") is not None else "",
            "Да" if rd.get("vak") else "",
            "Да" if rd.get("rinc") else "",
            "Да" if rd.get("rinc_core") else "",
            "Да" if rd.get("rsci") else "",
            "Да" if rd.get("white_list_flag") else "",
            wl_quartile,
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
