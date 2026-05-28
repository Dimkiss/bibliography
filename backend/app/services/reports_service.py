from __future__ import annotations

import io
import urllib.parse
from dataclasses import dataclass
from typing import Any

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.services.articles.bibliographic_reference import (
    build_bibliographic_reference,
)


SOURCE_TITLE_EXPR = """
    COALESCE(
        NULLIF(jn.JournalName, ''),
        NULLIF(j.jname, ''),
        NULLIF(a.Title_of_Material_F9, ''),
        NULLIF(a.Edition_F15, '')
    )
"""


@dataclass(frozen=True)
class ReportFile:
    content: bytes
    filename: str


def build_content_disposition(filename: str) -> str:
    ascii_name = (
        filename.encode("ascii", "ignore").decode("ascii").replace(" ", "_")
        or "report.xlsx"
    )
    encoded_name = urllib.parse.quote(filename, safe="")
    return (
        f'attachment; filename="{ascii_name}"; '
        f"filename*=UTF-8''{encoded_name}"
    )


def _yes(value: Any) -> str:
    return "Да" if value else ""


def _range_suffix(year_from: int | None, year_to: int | None) -> str:
    if year_from is not None and year_to is not None:
        return f"_{year_from}-{year_to}"
    if year_from is not None:
        return f"_from{year_from}"
    if year_to is not None:
        return f"_to{year_to}"
    return ""


def _range_title(year_from: int | None, year_to: int | None) -> str:
    if year_from is not None and year_to is not None:
        return f" ({year_from}-{year_to})"
    if year_from is not None:
        return f" (с {year_from})"
    if year_to is not None:
        return f" (по {year_to})"
    return ""


def _filename_token(value: str, fallback: str = "report") -> str:
    token = (
        value.replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
        .replace(" ", "_")
        .strip("_")
    )
    return token[:48] or fallback


def _make_in_filter(
    *,
    column: str,
    param_prefix: str,
    values: list[int],
    params: dict[str, Any],
) -> str:
    placeholders: list[str] = []
    for index, value in enumerate(values):
        key = f"{param_prefix}_{index}"
        params[key] = value
        placeholders.append(f":{key}")
    return f"{column} IN ({', '.join(placeholders)})"


def _fetch_author_columns(db: Session) -> set[str]:
    try:
        rows = db.execute(text("SHOW COLUMNS FROM authors")).mappings().all()
    except Exception:
        return {
            "authorID",
            "authorName",
            "position",
            "degree",
            "rank",
            "email",
            "WOS_ID",
            "Scopus_ID",
            "ORCID",
            "DepartmentCode",
        }

    return {str(row["Field"]) for row in rows}


def _style_workbook(
    *,
    title: str,
    sheet_name: str,
    headers: list[tuple[str, int]],
    rows: list[list[Any]],
) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="4E5B92")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    total_cols = len(headers)
    last_col = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 42

    for row_idx, values in enumerate(rows, start=3):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 48

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}{max(ws.max_row, 2)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _fetch_author_publication_rows(
    db: Session,
    *,
    author_ids: list[int],
    year_from: int | None = None,
    year_to: int | None = None,
    article_ids: list[int] | None = None,
) -> list[dict[str, Any]]:
    if not author_ids:
        return []

    params: dict[str, Any] = {}
    filters = [
        _make_in_filter(
            column="au.authorID",
            param_prefix="author_id",
            values=author_ids,
            params=params,
        )
    ]

    if year_from is not None:
        filters.append("a.Date_of_Publication_F20 >= :year_from")
        params["year_from"] = year_from
    if year_to is not None:
        filters.append("a.Date_of_Publication_F20 <= :year_to")
        params["year_to"] = year_to
    if article_ids:
        filters.append(
            _make_in_filter(
                column="a.Record_ID",
                param_prefix="article_id",
                values=article_ids,
                params=params,
            )
        )

    where_sql = " AND ".join(filters)

    sql = text(f"""
        SELECT
            au.authorID AS author_id,
            au.authorName AS author_name,
            au.position AS author_position,
            au.degree AS author_degree,
            au.rank AS author_rank,
            au.email AS author_email,
            au.WOS_ID AS author_wos_id,
            au.Scopus_ID AS author_scopus_id,
            au.ORCID AS author_orcid,
            au.DepartmentCode AS department_id,
            d.DepartmentName AS department_name,
            aha.AHA_ID AS author_link_id,
            aha.affiliation AS author_affiliation,
            COALESCE(aha.corresponding_author, 0) AS author_corresponding,
            CASE
                WHEN aha.AHA_ID = (
                    SELECT MIN(aha_first.AHA_ID)
                    FROM articlehasauthor aha_first
                    WHERE aha_first.Record_ID_f = a.Record_ID
                )
                THEN 1 ELSE 0
            END AS author_first,
            CASE
                WHEN aha.AHA_ID = (
                    SELECT MAX(aha_last.AHA_ID)
                    FROM articlehasauthor aha_last
                    WHERE aha_last.Record_ID_f = a.Record_ID
                )
                THEN 1 ELSE 0
            END AS author_last,
            a.Record_ID AS Record_ID,
            a.WorkFormType_f AS work_form_type,
            a.Author_Analitic_F1 AS author_analitic,
            a.Title_Analitic_F4 AS title,
            a.Title_Analitic_F4 AS title_analitic,
            a.Author_of_Material_F7 AS author_of_material,
            a.Title_of_Material_F9 AS title_of_material,
            a.DateOfMeeting_F12 AS date_of_meeting,
            a.Edition_F15 AS edition,
            a.Date_of_Publication_F20 AS date_of_publication,
            a.Date_of_Publication_F20 AS year,
            a.VolumeID_F22 AS volume,
            a.IssueID_F24 AS issue,
            a.Pages_F25 AS pages,
            a.ExtentOfWork_F26 AS extent_of_work,
            a.ISBN_F41 AS isbn,
            a.DOI AS doi,
            COALESCE(
                NULLIF(a.Author_Analitic_F1, ''),
                (
                    SELECT GROUP_CONCAT(au2.authorName ORDER BY aha2.AHA_ID SEPARATOR ', ')
                    FROM articlehasauthor aha2
                    JOIN authors au2 ON au2.authorID = aha2.authorID_f
                    WHERE aha2.Record_ID_f = a.Record_ID
                )
            ) AS authors,
            NULLIF(jn.JournalName, '') AS journal_name,
            {SOURCE_TITLE_EXPR} AS journal,
            jn.ISSN AS issn,
            pp.PlaceName AS place_name,
            pn.PublisherName AS publisher_name,
            NULLIF(j.Quartile, '') AS quartile,
            NULLIF(j.QuartileScopus, '') AS quartile_scopus,
            j.Impact_Factor AS impact_factor,
            COALESCE(j.WOS, 0) AS wos,
            COALESCE(j.Scopus, 0) AS scopus,
            COALESCE(j.LWL, 0) AS white_list_flag,
            j.LWL AS white_list_level,
            COALESCE(j.Rints, 0) AS rinc,
            COALESCE(j.RintsCore, 0) AS rinc_core,
            COALESCE(j.RSCI, 0) AS rsci,
            COALESCE(j.BAK, 0) AS vak,
            (
                SELECT GROUP_CONCAT(DISTINCT top.TOP_Name ORDER BY top.TOP_Name SEPARATOR ', ')
                FROM articlehastop aht
                JOIN typesofpublications top ON top.TOP_Flag = aht.TypeOfPublication_f
                WHERE aht.Record_ID_f = a.Record_ID
            ) AS pub_types
        FROM articlehasauthor aha
        JOIN authors au ON au.authorID = aha.authorID_f
        JOIN articles a ON a.Record_ID = aha.Record_ID_f
        LEFT JOIN departments d ON d.DepartmentCode = au.DepartmentCode
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        LEFT JOIN journalnames jn ON jn.JN_ID = j.JN_ID_f
        LEFT JOIN places pp ON pp.P_ID = a.PlaceOfPublication_F18_f
        LEFT JOIN publishernames pn ON pn.PN_ID = a.PublisherName_F19_f
        WHERE {where_sql}
        ORDER BY au.authorName ASC, a.Date_of_Publication_F20 DESC, a.Record_ID DESC
    """)

    return [dict(row) for row in db.execute(sql, params).mappings().all()]


PUBLICATION_HEADERS = [
    ("№", 5),
    ("Библиографическая\nссылка", 70),
    ("Название", 40),
    ("Авторы", 28),
    ("Журнал / Издание", 28),
    ("ISSN", 14),
    ("Год", 7),
    ("Том", 7),
    ("Выпуск", 8),
    ("Страницы", 11),
    ("DOI", 24),
    ("Тип публикации", 22),
    ("WoS", 7),
    ("Scopus", 8),
    ("Квартиль WoS", 13),
    ("Квартиль Scopus", 15),
    ("Импакт-\nфактор", 13),
    ("ВАК", 7),
    ("РИНЦ", 7),
    ("Ядро\nРИНЦ", 9),
    ("RSCI", 7),
    ("Белый\nсписок", 10),
    ("Уровень\nБС", 9),
]


def _publication_values(index: int, row: dict[str, Any]) -> list[Any]:
    wl_level = row.get("white_list_level")
    return [
        index,
        build_bibliographic_reference(row),
        row.get("title") or "",
        row.get("authors") or "",
        row.get("journal") or "",
        row.get("issn") or "",
        row.get("year"),
        row.get("volume") or "",
        row.get("issue") or "",
        row.get("pages") or "",
        row.get("doi") or "",
        row.get("pub_types") or "",
        _yes(row.get("wos")),
        _yes(row.get("scopus")),
        row.get("quartile") or "",
        row.get("quartile_scopus") or "",
        float(row["impact_factor"]) if row.get("impact_factor") is not None else "",
        _yes(row.get("vak")),
        _yes(row.get("rinc")),
        _yes(row.get("rinc_core")),
        _yes(row.get("rsci")),
        _yes(row.get("white_list_flag")),
        f"Q{wl_level}" if wl_level else "",
    ]


def generate_author_publications_report(
    db: Session,
    *,
    author_id: int,
    author_name: str,
    year_from: int | None = None,
    year_to: int | None = None,
    article_ids: list[int] | None = None,
) -> ReportFile:
    rows = _fetch_author_publication_rows(
        db,
        author_ids=[author_id],
        year_from=year_from,
        year_to=year_to,
        article_ids=article_ids,
    )
    data = [_publication_values(index, row) for index, row in enumerate(rows, start=1)]

    selected_suffix = f"_selected_{len(article_ids)}" if article_ids else ""
    range_suffix = "" if article_ids else _range_suffix(year_from, year_to)
    filename = (
        f"publications_{_filename_token(author_name, 'author')}"
        f"{selected_suffix or range_suffix}.xlsx"
    )

    return ReportFile(
        content=_style_workbook(
            title=f"Список публикаций: {author_name}{_range_title(year_from, year_to)}",
            sheet_name="Публикации",
            headers=PUBLICATION_HEADERS,
            rows=data,
        ),
        filename=filename,
    )


def _author_role_label(row: dict[str, Any]) -> str:
    roles: list[str] = []
    if row.get("author_first"):
        roles.append("Первый автор")
    if row.get("author_last"):
        roles.append("Последний автор")
    if row.get("author_corresponding"):
        roles.append("Автор для переписки")
    return ", ".join(roles) or "Соавтор"


AUTHORS_PUBLICATIONS_HEADERS = [
    ("№", 5),
    ("ID автора", 10),
    ("Автор", 32),
    ("Подразделение", 34),
    ("Должность", 24),
    ("Учёная степень", 22),
    ("Звание", 18),
    ("Email", 24),
    ("ORCID", 22),
    ("Scopus ID", 18),
    ("WOS ID", 18),
    ("Роль автора", 24),
    ("Affiliation", 11),
    ("Библиографическая\nссылка", 70),
    ("Название", 40),
    ("Авторы", 30),
    ("Журнал / Издание", 28),
    ("ISSN", 14),
    ("Год", 7),
    ("DOI", 24),
    ("Тип публикации", 22),
    ("WoS", 7),
    ("Scopus", 8),
    ("Квартиль WoS", 13),
    ("Квартиль Scopus", 15),
    ("Импакт-\nфактор", 13),
    ("ВАК", 7),
    ("РИНЦ", 7),
    ("Ядро\nРИНЦ", 9),
    ("RSCI", 7),
    ("Белый\nсписок", 10),
    ("Уровень\nБС", 9),
]


def _authors_publication_values(index: int, row: dict[str, Any]) -> list[Any]:
    publication_values = _publication_values(index, row)
    return [
        index,
        row.get("author_id"),
        row.get("author_name") or "",
        row.get("department_name") or "",
        row.get("author_position") or "",
        row.get("author_degree") or "",
        row.get("author_rank") or "",
        row.get("author_email") or "",
        row.get("author_orcid") or "",
        row.get("author_scopus_id") or "",
        row.get("author_wos_id") or "",
        _author_role_label(row),
        row.get("author_affiliation") or "",
        publication_values[1],
        publication_values[2],
        publication_values[3],
        publication_values[4],
        publication_values[5],
        publication_values[6],
        publication_values[10],
        publication_values[11],
        *publication_values[12:],
    ]


def generate_authors_publications_report(
    db: Session,
    *,
    author_ids: list[int],
    year_from: int | None = None,
    year_to: int | None = None,
    article_ids: list[int] | None = None,
) -> ReportFile:
    rows = _fetch_author_publication_rows(
        db,
        author_ids=author_ids,
        year_from=year_from,
        year_to=year_to,
        article_ids=article_ids,
    )
    data = [
        _authors_publication_values(index, row)
        for index, row in enumerate(rows, start=1)
    ]

    return ReportFile(
        content=_style_workbook(
            title=f"Публикации сотрудников{_range_title(year_from, year_to)}",
            sheet_name="Публикации сотрудников",
            headers=AUTHORS_PUBLICATIONS_HEADERS,
            rows=data,
        ),
        filename=f"authors_publications{_range_suffix(year_from, year_to)}.xlsx",
    )


def _fetch_author_summary_rows(
    db: Session,
    *,
    author_ids: list[int],
    year_from: int | None = None,
    year_to: int | None = None,
) -> list[dict[str, Any]]:
    if not author_ids:
        return []

    params: dict[str, Any] = {}
    author_filter = _make_in_filter(
        column="au.authorID",
        param_prefix="author_id",
        values=author_ids,
        params=params,
    )

    article_join_filters = ["a.Record_ID = aha.Record_ID_f"]
    if year_from is not None:
        article_join_filters.append("a.Date_of_Publication_F20 >= :year_from")
        params["year_from"] = year_from
    if year_to is not None:
        article_join_filters.append("a.Date_of_Publication_F20 <= :year_to")
        params["year_to"] = year_to
    article_join_sql = " AND ".join(article_join_filters)

    sql = text(f"""
        SELECT
            au.authorID AS author_id,
            au.authorName AS author_name,
            au.position AS author_position,
            au.degree AS author_degree,
            au.rank AS author_rank,
            au.email AS author_email,
            au.WOS_ID AS author_wos_id,
            au.Scopus_ID AS author_scopus_id,
            au.ORCID AS author_orcid,
            d.DepartmentName AS department_name,
            COUNT(DISTINCT a.Record_ID) AS total,
            COUNT(DISTINCT CASE WHEN COALESCE(j.WOS, 0) = 1 THEN a.Record_ID END)
                AS wos_count,
            COUNT(DISTINCT CASE WHEN COALESCE(j.Scopus, 0) = 1 THEN a.Record_ID END)
                AS scopus_count,
            COUNT(DISTINCT CASE WHEN COALESCE(j.LWL, 0) > 0 THEN a.Record_ID END)
                AS white_list_count,
            COUNT(DISTINCT CASE WHEN COALESCE(j.BAK, 0) = 1 THEN a.Record_ID END)
                AS vak_count,
            COUNT(DISTINCT CASE WHEN COALESCE(j.Rints, 0) = 1 THEN a.Record_ID END)
                AS rinc_count,
            COUNT(DISTINCT CASE WHEN COALESCE(j.RintsCore, 0) = 1 THEN a.Record_ID END)
                AS rinc_core_count,
            COUNT(DISTINCT CASE WHEN COALESCE(j.RSCI, 0) = 1 THEN a.Record_ID END)
                AS rsci_count,
            ROUND(SUM(COALESCE(j.Impact_Factor, 0)), 3) AS if_total,
            COUNT(DISTINCT CASE
                WHEN a.Record_ID IS NOT NULL
                    AND aha.AHA_ID = (
                        SELECT MIN(aha_first.AHA_ID)
                        FROM articlehasauthor aha_first
                        WHERE aha_first.Record_ID_f = a.Record_ID
                    )
                THEN a.Record_ID
            END) AS first_author_count,
            COUNT(DISTINCT CASE
                WHEN a.Record_ID IS NOT NULL
                    AND aha.AHA_ID = (
                        SELECT MAX(aha_last.AHA_ID)
                        FROM articlehasauthor aha_last
                        WHERE aha_last.Record_ID_f = a.Record_ID
                    )
                THEN a.Record_ID
            END) AS last_author_count,
            COUNT(DISTINCT CASE
                WHEN a.Record_ID IS NOT NULL
                    AND COALESCE(aha.corresponding_author, 0) = 1
                THEN a.Record_ID
            END) AS corresponding_author_count,
            MIN(a.Date_of_Publication_F20) AS first_year,
            MAX(a.Date_of_Publication_F20) AS last_year
        FROM authors au
        LEFT JOIN departments d ON d.DepartmentCode = au.DepartmentCode
        LEFT JOIN articlehasauthor aha ON aha.authorID_f = au.authorID
        LEFT JOIN articles a ON {article_join_sql}
        LEFT JOIN journals j ON j.J_ID = a.Journal_ID_f
        WHERE {author_filter}
        GROUP BY
            au.authorID,
            au.authorName,
            au.position,
            au.degree,
            au.rank,
            au.email,
            au.WOS_ID,
            au.Scopus_ID,
            au.ORCID,
            d.DepartmentName
        ORDER BY au.authorName ASC, au.authorID ASC
    """)

    return [dict(row) for row in db.execute(sql, params).mappings().all()]


AUTHORS_SUMMARY_HEADERS = [
    ("№", 5),
    ("ID автора", 10),
    ("Автор", 32),
    ("Подразделение", 34),
    ("Должность", 24),
    ("Учёная степень", 22),
    ("Звание", 18),
    ("Email", 24),
    ("ORCID", 22),
    ("Scopus ID", 18),
    ("WOS ID", 18),
    ("Всего", 9),
    ("WoS", 8),
    ("Scopus", 8),
    ("Белый список", 13),
    ("ВАК", 8),
    ("РИНЦ", 8),
    ("Ядро РИНЦ", 11),
    ("RSCI", 8),
    ("Суммарный ИФ", 13),
    ("Первым автором", 15),
    ("Последним автором", 17),
    ("Автор для переписки", 20),
    ("Первый год", 10),
    ("Последний год", 12),
]


def generate_authors_summary_report(
    db: Session,
    *,
    author_ids: list[int],
    year_from: int | None = None,
    year_to: int | None = None,
) -> ReportFile:
    rows = _fetch_author_summary_rows(
        db,
        author_ids=author_ids,
        year_from=year_from,
        year_to=year_to,
    )
    data = []
    for index, row in enumerate(rows, start=1):
        data.append(
            [
                index,
                row.get("author_id"),
                row.get("author_name") or "",
                row.get("department_name") or "",
                row.get("author_position") or "",
                row.get("author_degree") or "",
                row.get("author_rank") or "",
                row.get("author_email") or "",
                row.get("author_orcid") or "",
                row.get("author_scopus_id") or "",
                row.get("author_wos_id") or "",
                int(row.get("total") or 0),
                int(row.get("wos_count") or 0),
                int(row.get("scopus_count") or 0),
                int(row.get("white_list_count") or 0),
                int(row.get("vak_count") or 0),
                int(row.get("rinc_count") or 0),
                int(row.get("rinc_core_count") or 0),
                int(row.get("rsci_count") or 0),
                float(row.get("if_total") or 0),
                int(row.get("first_author_count") or 0),
                int(row.get("last_author_count") or 0),
                int(row.get("corresponding_author_count") or 0),
                row.get("first_year") or "",
                row.get("last_year") or "",
            ]
        )

    return ReportFile(
        content=_style_workbook(
            title=f"Сводка по сотрудникам{_range_title(year_from, year_to)}",
            sheet_name="Сводка",
            headers=AUTHORS_SUMMARY_HEADERS,
            rows=data,
        ),
        filename=f"authors_summary{_range_suffix(year_from, year_to)}.xlsx",
    )


def _format_author_status(value: Any) -> str:
    if value is None:
        return ""
    try:
        status = int(value)
    except (TypeError, ValueError):
        return str(value)
    return {
        0: "Уволен",
        1: "Работает",
        2: "Временно не работает",
    }.get(status, str(value))


def generate_authors_export_report(
    db: Session,
    *,
    author_ids: list[int],
) -> ReportFile:
    if not author_ids:
        return ReportFile(
            content=_style_workbook(
                title="Список авторов",
                sheet_name="Авторы",
                headers=[("№", 5)],
                rows=[],
            ),
            filename="authors.xlsx",
        )

    existing_columns = _fetch_author_columns(db)
    optional_selects: list[str] = []
    optional_headers: list[tuple[str, int, str]] = []

    optional_map = [
        ("type", "au.type AS author_type", "Тип", 9, "author_type"),
        ("status", "au.status AS author_status", "Статус", 18, "author_status"),
        ("birthdate", "au.birthdate AS birthdate", "Дата рождения", 14, "birthdate"),
        ("year", "au.year AS birth_year", "Год рождения", 13, "birth_year"),
        ("nickname", "au.nickname AS nickname", "Псевдоним", 24, "nickname"),
        ("Pattern", "au.Pattern AS search_pattern", "Шаблон поиска", 24, "search_pattern"),
        ("ID", "au.ID AS external_id", "Внешний ID", 12, "external_id"),
        (
            "snils",
            "RIGHT(au.snils, 4) AS snils_last4",
            "ID ПУ",
            10,
            "snils_last4",
        ),
    ]
    for column_name, select_sql, header, width, result_key in optional_map:
        if column_name in existing_columns:
            optional_selects.append(select_sql)
            optional_headers.append((header, width, result_key))

    params: dict[str, Any] = {}
    author_filter = _make_in_filter(
        column="au.authorID",
        param_prefix="author_id",
        values=author_ids,
        params=params,
    )
    optional_sql = ""
    if optional_selects:
        optional_sql = ",\n            " + ",\n            ".join(optional_selects)

    sql = text(f"""
        SELECT
            au.authorID AS author_id,
            au.authorName AS author_name,
            d.DepartmentName AS department_name,
            au.position AS author_position,
            au.degree AS author_degree,
            au.rank AS author_rank,
            au.email AS author_email,
            au.ORCID AS author_orcid,
            au.Scopus_ID AS author_scopus_id,
            au.WOS_ID AS author_wos_id
            {optional_sql}
        FROM authors au
        LEFT JOIN departments d ON d.DepartmentCode = au.DepartmentCode
        WHERE {author_filter}
        ORDER BY au.authorName ASC, au.authorID ASC
    """)
    rows = [dict(row) for row in db.execute(sql, params).mappings().all()]

    headers = [
        ("№", 5),
        ("ID автора", 10),
        ("Автор", 32),
        ("Подразделение", 34),
        ("Должность", 24),
        ("Учёная степень", 22),
        ("Звание", 18),
        ("Email", 24),
        ("ORCID", 22),
        ("Scopus ID", 18),
        ("WOS ID", 18),
        *[(header, width) for header, width, _ in optional_headers],
    ]

    data: list[list[Any]] = []
    for index, row in enumerate(rows, start=1):
        base_values = [
            index,
            row.get("author_id"),
            row.get("author_name") or "",
            row.get("department_name") or "",
            row.get("author_position") or "",
            row.get("author_degree") or "",
            row.get("author_rank") or "",
            row.get("author_email") or "",
            row.get("author_orcid") or "",
            row.get("author_scopus_id") or "",
            row.get("author_wos_id") or "",
        ]
        optional_values: list[Any] = []
        for _, _, result_key in optional_headers:
            value = row.get(result_key)
            if result_key == "author_status":
                optional_values.append(_format_author_status(value))
            else:
                optional_values.append(value or "")
        data.append([*base_values, *optional_values])

    return ReportFile(
        content=_style_workbook(
            title="Список авторов",
            sheet_name="Авторы",
            headers=headers,
            rows=data,
        ),
        filename="authors.xlsx",
    )
